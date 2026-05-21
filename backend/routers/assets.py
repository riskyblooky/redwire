from fastapi import APIRouter, Depends, HTTPException, status, Query, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from sqlalchemy.orm import selectinload
from typing import List, Optional
from database import get_db
from models.user import User
from models.asset import Asset
from models.asset_port import AssetPort, PortProtocol, PortState
from schemas.asset import AssetCreate, AssetUpdate, AssetResponse, AssetPortCreate, AssetPortUpdate, AssetPortResponse, AssetImportResult
from auth.dependencies import get_current_user
from auth.rbac import check_engagement_permission
from models.user import UserRole
from models.permission import Permission
import uuid
import io
import csv
import defusedxml.ElementTree as ET
from utils.collaboration import create_activity_log, build_change_summary
from models.discussion import ResourceType

router = APIRouter(prefix="/assets", tags=["assets"])


# ─── Template Downloads ──────────────────────────────────────────────────────

TEMPLATE_COLUMNS = ["name", "asset_type", "identifier", "description", "in_scope", "ports"]
TEMPLATE_EXAMPLE = ["Web Server", "SERVER", "192.168.1.10", "Main web server", "true", "80/tcp,443/tcp,22/tcp"]
ASSET_TYPES_HELP = "Valid asset_type values: IP_ADDRESS, DOMAIN, URL, APPLICATION, SERVER, NETWORK, OTHER"
PORTS_HELP = "Ports format: port/protocol (comma-separated). Example: 80/tcp,443/tcp,53/udp"


@router.get("/templates/csv")
async def download_csv_template():
    """Download a CSV template for asset import."""
    output = io.StringIO()
    writer = csv.writer(output)
    # Header comment row
    writer.writerow([f"# {ASSET_TYPES_HELP}"])
    writer.writerow([f"# {PORTS_HELP}"])
    writer.writerow(TEMPLATE_COLUMNS)
    writer.writerow(TEMPLATE_EXAMPLE)
    writer.writerow(["Corporate DNS", "DOMAIN", "example.com", "Primary domain", "true", "53/tcp,53/udp"])
    writer.writerow(["Login Portal", "URL", "https://example.com/login", "Customer login page", "true", ""])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=asset_import_template.csv"},
    )


@router.get("/templates/xlsx")
async def download_xlsx_template():
    """Download an XLSX template for asset import."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Assets"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    # Column widths
    col_widths = {"A": 25, "B": 18, "C": 35, "D": 40, "E": 12, "F": 30}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # Headers
    for col_idx, header in enumerate(TEMPLATE_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Example rows
    examples = [
        TEMPLATE_EXAMPLE,
        ["Corporate DNS", "DOMAIN", "example.com", "Primary domain", "true", "53/tcp,53/udp"],
        ["Login Portal", "URL", "https://example.com/login", "Customer login page", "true", ""],
        ["Internal App", "APPLICATION", "ERP System", "SAP instance", "true", "8080/tcp"],
    ]
    for row_idx, example in enumerate(examples, 2):
        for col_idx, value in enumerate(example, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    # Help sheet
    help_ws = wb.create_sheet("Help")
    help_ws.column_dimensions["A"].width = 20
    help_ws.column_dimensions["B"].width = 60
    help_data = [
        ("Column", "Description"),
        ("name", "Display name for the asset (required)"),
        ("asset_type", ASSET_TYPES_HELP),
        ("identifier", "IP address, domain, URL, or other identifier (required)"),
        ("description", "Optional description"),
        ("in_scope", "'true' or 'false' (default: true)"),
        ("ports", PORTS_HELP),
    ]
    for row_idx, (col_a, col_b) in enumerate(help_data, 1):
        cell_a = help_ws.cell(row=row_idx, column=1, value=col_a)
        cell_b = help_ws.cell(row=row_idx, column=2, value=col_b)
        if row_idx == 1:
            cell_a.font = header_font
            cell_a.fill = header_fill
            cell_b.font = header_font
            cell_b.fill = header_fill

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=asset_import_template.xlsx"},
    )


# ─── Asset Import ─────────────────────────────────────────────────────────────

def _parse_ports(ports_str: str) -> list:
    """Parse a ports string like '80/tcp,443/tcp,53/udp' into structured port dicts."""
    if not ports_str or not ports_str.strip():
        return []
    ports = []
    for part in ports_str.split(","):
        part = part.strip()
        if not part:
            continue
        if "/" in part:
            port_num, proto = part.split("/", 1)
        else:
            port_num = part
            proto = "tcp"
        try:
            port_int = int(port_num.strip())
            if port_int < 1 or port_int > 65535:
                continue
            proto_upper = proto.strip().upper()
            protocol = PortProtocol.TCP if proto_upper != "UDP" else PortProtocol.UDP
            ports.append({"port_number": port_int, "protocol": protocol})
        except ValueError:
            continue
    return ports


def _resolve_asset_type(type_str: str) -> Optional[str]:
    """Resolve an asset type string to a canonical name, with fuzzy matching."""
    type_str = type_str.strip().upper().replace(" ", "_")
    mapping = {
        "IP_ADDRESS": "IP Address", "IP": "IP Address", "IPADDRESS": "IP Address",
        "DOMAIN": "Domain", "DNS": "Domain",
        "URL": "URL",
        "APPLICATION": "Application", "APP": "Application",
        "SERVER": "Server", "HOST": "Server",
        "NETWORK": "Network", "SUBNET": "Network",
        "OTHER": "Other",
    }
    return mapping.get(type_str)


def _parse_csv_rows(content: bytes) -> list:
    """Parse CSV content into list of row dicts."""
    text = content.decode("utf-8-sig")  # handle BOM
    reader = csv.DictReader(
        (line for line in text.splitlines() if not line.strip().startswith("#")),
        fieldnames=None,
    )
    rows = []
    for row in reader:
        # Normalize keys to lowercase, strip whitespace
        normalized = {k.strip().lower(): (v.strip() if v else "") for k, v in row.items() if k}
        rows.append(normalized)
    return rows


def _parse_xlsx_rows(content: bytes) -> list:
    """Parse XLSX content into list of row dicts."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    # First row = headers
    headers = [str(h).strip().lower() if h else "" for h in rows[0]]
    result = []
    for row in rows[1:]:
        if all(cell is None for cell in row):
            continue
        row_dict = {}
        for idx, header in enumerate(headers):
            if header and idx < len(row):
                val = row[idx]
                row_dict[header] = str(val).strip() if val is not None else ""
        result.append(row_dict)
    return result


def _parse_nmap_xml(content: bytes) -> list:
    """Parse NMAP XML output into asset + port structures."""
    assets = []
    try:
        root = ET.fromstring(content)
    except ET.ParseError:
        return []

    for host in root.findall(".//host"):
        # Skip hosts that are down
        status_elem = host.find("status")
        if status_elem is not None and status_elem.get("state") != "up":
            continue

        # Get address
        addr_elem = host.find("address")
        if addr_elem is None:
            continue

        addr = addr_elem.get("addr", "")
        addr_type = addr_elem.get("addrtype", "ipv4")

        if not addr:
            continue

        # Determine asset type
        if addr_type in ("ipv4", "ipv6"):
            asset_type = "IP Address"
        else:
            asset_type = "Other"

        # Try to get hostname
        hostname = None
        hostnames = host.find("hostnames")
        if hostnames is not None:
            hn = hostnames.find("hostname")
            if hn is not None:
                hostname = hn.get("name")

        name = hostname or addr

        # Parse ports
        ports = []
        ports_elem = host.find("ports")
        if ports_elem is not None:
            for port_elem in ports_elem.findall("port"):
                port_num = port_elem.get("portid")
                protocol = port_elem.get("protocol", "tcp").upper()

                state_elem = port_elem.find("state")
                state = "OPEN"
                if state_elem is not None:
                    state_val = state_elem.get("state", "open").upper()
                    if state_val in ("OPEN", "CLOSED", "FILTERED"):
                        state = state_val

                service_elem = port_elem.find("service")
                service_name = None
                version = None
                if service_elem is not None:
                    service_name = service_elem.get("name")
                    product = service_elem.get("product", "")
                    svc_version = service_elem.get("version", "")
                    if product or svc_version:
                        version = f"{product} {svc_version}".strip()

                try:
                    port_int = int(port_num)
                    if 1 <= port_int <= 65535:
                        port_protocol = PortProtocol.UDP if protocol == "UDP" else PortProtocol.TCP
                        port_state = PortState.OPEN
                        if state == "CLOSED":
                            port_state = PortState.CLOSED
                        elif state == "FILTERED":
                            port_state = PortState.FILTERED

                        ports.append({
                            "port_number": port_int,
                            "protocol": port_protocol,
                            "service_name": service_name,
                            "state": port_state,
                            "version": version,
                        })
                except (ValueError, TypeError):
                    continue

        assets.append({
            "name": name,
            "asset_type": asset_type,
            "identifier": addr,
            "description": f"Imported from NMAP scan{' (' + hostname + ')' if hostname and hostname != addr else ''}",
            "ports": ports,
        })

    return assets


@router.post("/import", response_model=AssetImportResult)
async def import_assets(
    file: UploadFile = File(...),
    engagement_id: str = Form(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Import assets from CSV, XLSX, or NMAP XML file."""
    # Check permissions
    is_admin = current_user.role in [UserRole.ADMIN, UserRole.READ_ONLY_ADMIN, UserRole.TEAM_LEAD]
    if not is_admin:
        has_permission = await check_engagement_permission(
            current_user.id, engagement_id, Permission.ASSET_CREATE.value, db
        )
        if not has_permission:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Insufficient permissions to import assets.",
            )

    content = await file.read()
    original_filename = file.filename or "unknown"
    filename = original_filename.lower()

    # Detect format and parse
    parsed_assets = []
    is_nmap = False

    if filename.endswith(".csv"):
        rows = _parse_csv_rows(content)
        for row in rows:
            asset_type = _resolve_asset_type(row.get("asset_type", "OTHER"))
            if not asset_type:
                asset_type = "Other"
            in_scope = row.get("in_scope", "true").lower() not in ("false", "0", "no", "n")
            parsed_assets.append({
                "name": row.get("name", "").strip(),
                "asset_type": asset_type,
                "identifier": row.get("identifier", "").strip(),
                "description": row.get("description", ""),
                "in_scope": in_scope,
                "ports": _parse_ports(row.get("ports", "")),
            })
    elif filename.endswith(".xlsx") or filename.endswith(".xls"):
        rows = _parse_xlsx_rows(content)
        for row in rows:
            asset_type = _resolve_asset_type(row.get("asset_type", "OTHER"))
            if not asset_type:
                asset_type = "Other"
            in_scope = row.get("in_scope", "true").lower() not in ("false", "0", "no", "n")
            parsed_assets.append({
                "name": row.get("name", "").strip(),
                "asset_type": asset_type,
                "identifier": row.get("identifier", "").strip(),
                "description": row.get("description", ""),
                "in_scope": in_scope,
                "ports": _parse_ports(row.get("ports", "")),
            })
    elif filename.endswith(".xml"):
        parsed_assets = _parse_nmap_xml(content)
        is_nmap = True
    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Unsupported file format. Use .csv, .xlsx, or .xml (NMAP).",
        )

    if not parsed_assets:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No assets found in the uploaded file.",
        )

    # Get existing assets for dedup
    existing_result = await db.execute(
        select(Asset).where(Asset.engagement_id == engagement_id)
    )
    existing_assets = {a.identifier.lower(): a for a in existing_result.scalars().all()}

    result = AssetImportResult()

    for asset_data in parsed_assets:
        name = asset_data.get("name", "")
        identifier = asset_data.get("identifier", "")

        if not name or not identifier:
            result.errors.append(f"Skipped row: missing name or identifier")
            continue

        identifier_lower = identifier.lower()

        if identifier_lower in existing_assets:
            # Merge ports for existing assets
            existing_asset = existing_assets[identifier_lower]
            existing_ports = {(p.port_number, p.protocol) for p in existing_asset.ports}

            for port_data in asset_data.get("ports", []):
                port_key = (port_data["port_number"], port_data["protocol"])
                if port_key not in existing_ports:
                    db_port = AssetPort(
                        id=str(uuid.uuid4()),
                        asset_id=existing_asset.id,
                        port_number=port_data["port_number"],
                        protocol=port_data["protocol"],
                        service_name=port_data.get("service_name"),
                        state=port_data.get("state", PortState.OPEN),
                        version=port_data.get("version"),
                    )
                    db.add(db_port)
                    existing_ports.add(port_key)
                    result.ports_added += 1

            result.skipped += 1
            continue

        # Create new asset
        db_asset = Asset(
            id=str(uuid.uuid4()),
            engagement_id=engagement_id,
            name=name,
            asset_type=asset_data.get("asset_type", "Other"),
            identifier=identifier,
            description=f"Imported from: {original_filename}" + (f"\n\n{asset_data['description']}" if asset_data.get('description') else ""),
            in_scope=asset_data.get("in_scope", True),
            is_scanned=is_nmap,  # Mark as scanned if from NMAP
            created_by=current_user.id,
        )
        db.add(db_asset)

        # Add ports
        for port_data in asset_data.get("ports", []):
            db_port = AssetPort(
                id=str(uuid.uuid4()),
                asset_id=db_asset.id,
                port_number=port_data["port_number"],
                protocol=port_data["protocol"],
                service_name=port_data.get("service_name"),
                state=port_data.get("state", PortState.OPEN),
                version=port_data.get("version"),
            )
            db.add(db_port)
            result.ports_added += 1

        existing_assets[identifier_lower] = db_asset
        result.created += 1

    await db.commit()

    # Log activity
    source = "NMAP XML" if is_nmap else filename.rsplit(".", 1)[-1].upper()
    await create_activity_log(
        db,
        engagement_id=engagement_id,
        user_id=current_user.id,
        action="imported_assets",
        resource_type="asset",
        resource_id="",
        resource_name="Bulk Import",
        details=f"Imported assets from {source}: {result.created} created, {result.skipped} skipped, {result.ports_added} ports added",
    )

    return result


# ─── Port CRUD ────────────────────────────────────────────────────────────────

@router.post("/{asset_id}/ports", response_model=AssetPortResponse, status_code=status.HTTP_201_CREATED)
async def add_asset_port(
    asset_id: str,
    port_data: AssetPortCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Add a port to an asset."""
    result = await db.execute(select(Asset).where(Asset.id == asset_id))
    asset = result.scalar_one_or_none()
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")

    # Permission check
    is_admin = current_user.role in [UserRole.ADMIN, UserRole.READ_ONLY_ADMIN, UserRole.TEAM_LEAD]
    if not is_admin:
        has_permission = await check_engagement_permission(
            current_user.id, asset.engagement_id, Permission.ASSET_EDIT.value, db
        )
        if not has_permission:
            raise HTTPException(status_code=403, detail="Insufficient permissions")

    # Check for duplicates
    existing = await db.execute(
        select(AssetPort).where(
            AssetPort.asset_id == asset_id,
            AssetPort.port_number == port_data.port_number,
            AssetPort.protocol == port_data.protocol,
        )
    )
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=409, detail="Port already exists on this asset")

    db_port = AssetPort(
        id=str(uuid.uuid4()),
        asset_id=asset_id,
        **port_data.model_dump(),
    )
    db.add(db_port)
    await db.commit()
    await db.refresh(db_port)

    return db_port


@router.put("/{asset_id}/ports/{port_id}", response_model=AssetPortResponse)
async def update_asset_port(
    asset_id: str,
    port_id: str,
    port_data: AssetPortUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Update a port on an asset."""

    result = await db.execute(
        select(AssetPort).where(AssetPort.id == port_id, AssetPort.asset_id == asset_id)
    )
    db_port = result.scalar_one_or_none()
    if not db_port:
        raise HTTPException(status_code=404, detail="Port not found")

    # Permission check
    asset_result = await db.execute(select(Asset).where(Asset.id == asset_id))
    asset = asset_result.scalar_one_or_none()
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")

    is_admin = current_user.role in [UserRole.ADMIN, UserRole.READ_ONLY_ADMIN, UserRole.TEAM_LEAD]
    if not is_admin:
        has_permission = await check_engagement_permission(
            current_user.id, asset.engagement_id, Permission.ASSET_EDIT.value, db
        )
        if not has_permission:
            raise HTTPException(status_code=403, detail="Insufficient permissions")

    update_data = port_data.model_dump(exclude_unset=True)

    # Check for duplicate if port_number or protocol is changing
    new_port_num = update_data.get("port_number", db_port.port_number)
    new_protocol = update_data.get("protocol", db_port.protocol)
    if new_port_num != db_port.port_number or new_protocol != db_port.protocol:
        existing = await db.execute(
            select(AssetPort).where(
                AssetPort.asset_id == asset_id,
                AssetPort.port_number == new_port_num,
                AssetPort.protocol == new_protocol,
                AssetPort.id != port_id,
            )
        )
        if existing.scalar_one_or_none():
            raise HTTPException(status_code=409, detail="Port already exists on this asset")

    for key, value in update_data.items():
        setattr(db_port, key, value)

    await db.commit()
    await db.refresh(db_port)
    return db_port


@router.delete("/{asset_id}/ports/{port_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_asset_port(
    asset_id: str,
    port_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Remove a port from an asset."""
    result = await db.execute(
        select(AssetPort).where(AssetPort.id == port_id, AssetPort.asset_id == asset_id)
    )
    db_port = result.scalar_one_or_none()
    if not db_port:
        raise HTTPException(status_code=404, detail="Port not found")

    # Get asset for permission check
    asset_result = await db.execute(select(Asset).where(Asset.id == asset_id))
    asset = asset_result.scalar_one_or_none()
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")

    is_admin = current_user.role in [UserRole.ADMIN, UserRole.READ_ONLY_ADMIN, UserRole.TEAM_LEAD]
    if not is_admin:
        has_permission = await check_engagement_permission(
            current_user.id, asset.engagement_id, Permission.ASSET_EDIT.value, db
        )
        if not has_permission:
            raise HTTPException(status_code=403, detail="Insufficient permissions")

    await db.delete(db_port)
    await db.commit()
    return None


# ─── Existing CRUD Endpoints ─────────────────────────────────────────────────


@router.get("/port-filters")
async def get_port_filters(
    engagement_id: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get distinct port numbers and service names for filter dropdowns"""
    from sqlalchemy import func, distinct

    port_query = select(
        distinct(AssetPort.port_number),
        AssetPort.protocol,
    ).join(Asset, AssetPort.asset_id == Asset.id).where(
        AssetPort.state == PortState.OPEN
    )

    service_query = select(
        distinct(AssetPort.service_name)
    ).join(Asset, AssetPort.asset_id == Asset.id).where(
        AssetPort.state == PortState.OPEN,
        AssetPort.service_name.isnot(None),
        AssetPort.service_name != '',
    )

    if engagement_id:
        port_query = port_query.where(Asset.engagement_id == engagement_id)
        service_query = service_query.where(Asset.engagement_id == engagement_id)

    port_result = await db.execute(port_query.order_by(AssetPort.port_number))
    service_result = await db.execute(service_query.order_by(AssetPort.service_name))

    ports = [{"port_number": row[0], "protocol": row[1].value if row[1] else "TCP"} for row in port_result.all()]
    services = [row[0] for row in service_result.all()]

    return {"ports": ports, "services": services}

@router.get("")
async def get_assets(
    engagement_id: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    port: Optional[int] = Query(None, description="Filter assets that have this port number"),
    service: Optional[str] = Query(None, description="Filter assets that have this service name"),
    port_state: Optional[str] = Query(None, description="Port state filter (OPEN, CLOSED, FILTERED). Defaults to OPEN when port/service filters active"),
    sort_by: Optional[str] = Query("created_at"),
    sort_order: Optional[str] = Query("desc"),
    skip: int = 0,
    limit: int = 100,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get all assets, optionally filtered by engagement"""
    from models.discussion import Thread
    from sqlalchemy import func
    
    # Join with threads to count unresolved threads
    query = select(
        Asset,
        User.username.label("creator_username"),
        User.profile_photo.label("creator_profile_photo"),
        func.count(Thread.id).filter(Thread.is_resolved == False).label("unresolved_count")
    ).outerjoin(
        Thread,
        (Thread.resource_type == "asset") & (Thread.resource_id == Asset.id)
    ).outerjoin(
        User,
        Asset.created_by == User.id
    ).group_by(Asset.id, User.username, User.profile_photo)
    
    if engagement_id:
        query = query.where(Asset.engagement_id == engagement_id)
    
    if search:
        search_term = f"%{search}%"
        from sqlalchemy import or_
        query = query.where(
            or_(
                Asset.name.ilike(search_term),
                Asset.identifier.ilike(search_term),
                Asset.description.ilike(search_term),
            )
        )
    
    # Port / service filtering
    if port is not None or service:
        query = query.join(AssetPort, AssetPort.asset_id == Asset.id)
        if port is not None:
            query = query.where(AssetPort.port_number == port)
        if service:
            query = query.where(AssetPort.service_name.ilike(f"%{service}%"))
        if port_state:
            query = query.where(AssetPort.state == port_state)
        else:
            query = query.where(AssetPort.state == PortState.OPEN)
    
    # Restrict to assigned engagements for non-admins
    if current_user.role not in [UserRole.ADMIN, UserRole.READ_ONLY_ADMIN, UserRole.TEAM_LEAD]:
        from models.engagement import Engagement
        query = query.join(Engagement, Asset.engagement_id == Engagement.id).where(
            Engagement.assigned_users.any(User.id == current_user.id)
        )
    
    # Get total count before pagination
    count_query = select(func.count()).select_from(query.subquery())
    total_result = await db.execute(count_query)
    total = total_result.scalar() or 0

    # Sorting
    sort_column = Asset.created_at
    if sort_by == 'name':
        sort_column = Asset.name
    elif sort_by == 'asset_type':
        sort_column = Asset.asset_type

    # When searching, sort by relevance first (best match first)
    if search:
        from sqlalchemy import case
        relevance = case(
            (Asset.name.ilike(search), 0),                    # exact name match
            (Asset.name.ilike(f"{search}%"), 1),              # name starts with
            (Asset.identifier.ilike(search), 2),              # exact identifier match
            (Asset.identifier.ilike(f"{search}%"), 3),        # identifier starts with
            else_=4                                           # contains somewhere
        )
        if sort_order == 'asc':
            query = query.order_by(relevance.asc(), sort_column.asc())
        else:
            query = query.order_by(relevance.asc(), sort_column.desc())
    elif sort_order == 'asc':
        query = query.order_by(sort_column.asc())
    else:
        query = query.order_by(sort_column.desc())
    
    query = query.offset(skip).limit(limit)
    
    result = await db.execute(query)
    assets_with_counts = []
    for asset, creator_username, creator_profile_photo, unresolved_count in result.all():
        asset_dict = AssetResponse.model_validate(asset).model_dump()
        asset_dict["unresolved_thread_count"] = unresolved_count or 0
        asset_dict["created_by_username"] = creator_username
        asset_dict["created_by_profile_photo"] = creator_profile_photo
        assets_with_counts.append(asset_dict)
    
    return {"items": assets_with_counts, "total": total}

@router.get("/{asset_id}", response_model=AssetResponse)
async def get_asset(
    asset_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get a specific asset by ID"""
    result = await db.execute(
        select(Asset, User.username, User.profile_photo)
        .options(
            selectinload(Asset.findings),
            selectinload(Asset.testcases),
            selectinload(Asset.vault_items),
            selectinload(Asset.cleanup_artifacts),
            selectinload(Asset.ports),
        )
        .outerjoin(User, Asset.created_by == User.id)
        .where(Asset.id == asset_id)
    )
    row = result.first()
    
    if not row:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Asset not found"
        )
    
    asset, creator_username, creator_profile_photo = row
    
    # Authorization Check using RBAC
    is_admin = current_user.role in [UserRole.ADMIN, UserRole.READ_ONLY_ADMIN, UserRole.TEAM_LEAD]
    if not is_admin:
        has_permission = await check_engagement_permission(current_user.id, asset.engagement_id, Permission.ASSET_VIEW.value, db)
        if not has_permission:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Insufficient permissions. You need the 'asset_view' permission to view assets."
            )
    
    asset_dict = AssetResponse.model_validate(asset).model_dump()
    asset_dict["created_by_username"] = creator_username
    asset_dict["created_by_profile_photo"] = creator_profile_photo
    return AssetResponse(**asset_dict)

@router.post("", response_model=AssetResponse, status_code=status.HTTP_201_CREATED)
async def create_asset(
    asset_data: AssetCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Create a new asset"""
    # Check permissions using RBAC
    is_admin = current_user.role in [UserRole.ADMIN, UserRole.READ_ONLY_ADMIN, UserRole.TEAM_LEAD]
    if not is_admin:
        has_permission = await check_engagement_permission(current_user.id, asset_data.engagement_id, Permission.ASSET_CREATE.value, db)
        if not has_permission:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Insufficient permissions. You need the 'asset_create' permission to add assets."
            )

    db_asset = Asset(
        id=str(uuid.uuid4()),
        created_by=current_user.id,
        **asset_data.model_dump()
    )
    
    db.add(db_asset)
    await db.commit()
    await db.refresh(db_asset)
    
    # Log activity
    await create_activity_log(
        db,
        engagement_id=asset_data.engagement_id,
        user_id=current_user.id,
        action="created_asset",
        resource_type="asset",
        resource_id=db_asset.id,
        resource_name=db_asset.name,
        details=f"Created asset: {db_asset.name} ({db_asset.asset_type})",
        extra_context={
            "asset_type": str(db_asset.asset_type).lower() if db_asset.asset_type else None,
        },
    )

    return db_asset

@router.put("/{asset_id}", response_model=AssetResponse)
async def update_asset(
    asset_id: str,
    asset_update: AssetUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Update an existing asset"""
    result = await db.execute(select(Asset).where(Asset.id == asset_id))
    db_asset = result.scalar_one_or_none()
    
    if not db_asset:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Asset not found"
        )
    
    # Check permissions using RBAC with ANY model
    is_admin = current_user.role in [UserRole.ADMIN, UserRole.READ_ONLY_ADMIN, UserRole.TEAM_LEAD]
    is_owner = db_asset.created_by == current_user.id
    
    if not is_admin:
        if is_owner:
            # Owner needs base edit permission
            has_permission = await check_engagement_permission(current_user.id, db_asset.engagement_id, Permission.ASSET_EDIT.value, db)
        else:
            # Non-owner needs edit_any permission
            has_permission = await check_engagement_permission(current_user.id, db_asset.engagement_id, Permission.ASSET_EDIT_ANY.value, db)
        
        if not has_permission:
            required_perm = Permission.ASSET_EDIT.value if is_owner else Permission.ASSET_EDIT_ANY.value
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail=f"Insufficient permissions. You need the '{required_perm}' permission to modify this asset."
            )
    
    # Update fields
    update_data = asset_update.model_dump(exclude_unset=True)

    # Capture change summary before applying updates
    change_details = build_change_summary(db_asset, update_data, label=f"Updated asset '{db_asset.name}'")

    for field, value in update_data.items():
        setattr(db_asset, field, value)
    
    db_asset.updated_by = current_user.id
    
    await db.commit()
    await db.refresh(db_asset)
    
    # Log activity
    await create_activity_log(
        db,
        engagement_id=db_asset.engagement_id,
        user_id=current_user.id,
        action="updated_asset",
        resource_type="asset",
        resource_id=db_asset.id,
        resource_name=db_asset.name,
        details=change_details,
        extra_context={
            "asset_type": str(db_asset.asset_type).lower() if db_asset.asset_type else None,
        },
    )

    return db_asset

@router.delete("/{asset_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_asset(
    asset_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Delete an asset"""
    result = await db.execute(select(Asset).where(Asset.id == asset_id))
    db_asset = result.scalar_one_or_none()
    
    if not db_asset:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Asset not found"
        )
    
    # Check permissions using RBAC with ANY model
    is_admin = current_user.role in [UserRole.ADMIN, UserRole.READ_ONLY_ADMIN, UserRole.TEAM_LEAD]
    is_owner = db_asset.created_by == current_user.id
    
    if not is_admin:
        if is_owner:
            # Owner needs base delete permission
            has_permission = await check_engagement_permission(current_user.id, db_asset.engagement_id, Permission.ASSET_DELETE.value, db)
        else:
            # Non-owner needs delete_any permission
            has_permission = await check_engagement_permission(current_user.id, db_asset.engagement_id, Permission.ASSET_DELETE_ANY.value, db)
        
        if not has_permission:
            required_perm = Permission.ASSET_DELETE.value if is_owner else Permission.ASSET_DELETE_ANY.value
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail=f"Insufficient permissions. You need the '{required_perm}' permission to delete this asset."
            )
    
    # Log activity before deletion
    await create_activity_log(
        db,
        engagement_id=db_asset.engagement_id,
        user_id=current_user.id,
        action="deleted_asset",
        resource_type="asset",
        resource_id=db_asset.id,
        resource_name=db_asset.name,
        details=f"Deleted asset: {db_asset.name}"
    )

    await db.delete(db_asset)
    await db.commit()
    
    return None


# ── Cross-Link Endpoints ──────────────────────────────────────────────────────

async def _require_asset(asset_id: str, db: AsyncSession, current_user: User) -> Asset:
    """Load an asset and verify the user has edit permission."""
    result = await db.execute(select(Asset).where(Asset.id == asset_id))
    asset = result.scalar_one_or_none()
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")
    is_admin = current_user.role in [UserRole.ADMIN, UserRole.READ_ONLY_ADMIN, UserRole.TEAM_LEAD]
    if not is_admin:
        has_perm = await check_engagement_permission(
            current_user.id, asset.engagement_id, Permission.ASSET_EDIT.value, db
        )
        if not has_perm:
            raise HTTPException(status_code=403, detail="Insufficient permissions.")
    return asset


# ── Asset ↔ Finding ──

@router.post("/{asset_id}/findings/{finding_id}", status_code=status.HTTP_204_NO_CONTENT)
async def link_asset_to_finding(asset_id: str, finding_id: str, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    """Link an asset to a finding (without remediation fields)."""
    from models.associations import FindingAsset
    from models.finding import Finding
    asset = await _require_asset(asset_id, db, current_user)
    existing = await db.execute(select(FindingAsset).where(FindingAsset.asset_id == asset_id, FindingAsset.finding_id == finding_id))
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=409, detail="Already linked")
    finding = (await db.execute(select(Finding).where(Finding.id == finding_id))).scalar_one_or_none()
    if not finding:
        raise HTTPException(status_code=404, detail="Finding not found")
    if finding.engagement_id != asset.engagement_id:
        raise HTTPException(status_code=400, detail="Finding belongs to a different engagement")
    db.add(FindingAsset(finding_id=finding_id, asset_id=asset_id))
    await db.commit()


@router.delete("/{asset_id}/findings/{finding_id}", status_code=status.HTTP_204_NO_CONTENT)
async def unlink_asset_from_finding(asset_id: str, finding_id: str, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    """Remove the link between an asset and a finding."""
    from models.associations import FindingAsset
    await _require_asset(asset_id, db, current_user)
    result = await db.execute(select(FindingAsset).where(FindingAsset.asset_id == asset_id, FindingAsset.finding_id == finding_id))
    link = result.scalar_one_or_none()
    if not link:
        raise HTTPException(status_code=404, detail="Link not found")
    await db.delete(link)
    await db.commit()


# ── Asset ↔ TestCase ──

@router.post("/{asset_id}/testcases/{testcase_id}", status_code=status.HTTP_204_NO_CONTENT)
async def link_asset_to_testcase(asset_id: str, testcase_id: str, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    """Link an asset to a test case."""
    from models.associations import TestCaseAsset
    from models.testcase import TestCase as TC
    asset = await _require_asset(asset_id, db, current_user)
    existing = await db.execute(select(TestCaseAsset).where(TestCaseAsset.asset_id == asset_id, TestCaseAsset.testcase_id == testcase_id))
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=409, detail="Already linked")
    tc = (await db.execute(select(TC).where(TC.id == testcase_id))).scalar_one_or_none()
    if not tc:
        raise HTTPException(status_code=404, detail="Test case not found")
    if tc.engagement_id != asset.engagement_id:
        raise HTTPException(status_code=400, detail="Test case belongs to a different engagement")
    db.add(TestCaseAsset(testcase_id=testcase_id, asset_id=asset_id))
    await db.commit()


@router.delete("/{asset_id}/testcases/{testcase_id}", status_code=status.HTTP_204_NO_CONTENT)
async def unlink_asset_from_testcase(asset_id: str, testcase_id: str, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    """Remove the link between an asset and a test case."""
    from models.associations import TestCaseAsset
    await _require_asset(asset_id, db, current_user)
    result = await db.execute(select(TestCaseAsset).where(TestCaseAsset.asset_id == asset_id, TestCaseAsset.testcase_id == testcase_id))
    link = result.scalar_one_or_none()
    if not link:
        raise HTTPException(status_code=404, detail="Link not found")
    await db.delete(link)
    await db.commit()


# ── Asset ↔ VaultItem ──

@router.post("/{asset_id}/vault-items/{vault_item_id}", status_code=status.HTTP_204_NO_CONTENT)
async def link_asset_to_vault_item(asset_id: str, vault_item_id: str, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    """Link an asset to a vault item."""
    from models.associations import VaultItemAsset
    from models.vault import VaultItem
    asset = await _require_asset(asset_id, db, current_user)
    existing = await db.execute(select(VaultItemAsset).where(VaultItemAsset.asset_id == asset_id, VaultItemAsset.vault_item_id == vault_item_id))
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=409, detail="Already linked")
    item = (await db.execute(select(VaultItem).where(VaultItem.id == vault_item_id))).scalar_one_or_none()
    if not item:
        raise HTTPException(status_code=404, detail="Vault item not found")
    if item.engagement_id != asset.engagement_id:
        raise HTTPException(status_code=400, detail="Vault item belongs to a different engagement")
    db.add(VaultItemAsset(vault_item_id=vault_item_id, asset_id=asset_id))
    await db.commit()


@router.delete("/{asset_id}/vault-items/{vault_item_id}", status_code=status.HTTP_204_NO_CONTENT)
async def unlink_asset_from_vault_item(asset_id: str, vault_item_id: str, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    """Remove the link between an asset and a vault item."""
    from models.associations import VaultItemAsset
    await _require_asset(asset_id, db, current_user)
    result = await db.execute(select(VaultItemAsset).where(VaultItemAsset.asset_id == asset_id, VaultItemAsset.vault_item_id == vault_item_id))
    link = result.scalar_one_or_none()
    if not link:
        raise HTTPException(status_code=404, detail="Link not found")
    await db.delete(link)
    await db.commit()


@router.post("/{asset_id}/cleanup-artifacts/{cleanup_artifact_id}", status_code=status.HTTP_204_NO_CONTENT)
async def link_asset_to_cleanup_artifact(asset_id: str, cleanup_artifact_id: str, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    """Link an asset to a cleanup artifact."""
    from models.associations import CleanupArtifactAsset
    from models.cleanup_artifact import CleanupArtifact as CA
    asset = await _require_asset(asset_id, db, current_user)
    existing = await db.execute(select(CleanupArtifactAsset).where(CleanupArtifactAsset.asset_id == asset_id, CleanupArtifactAsset.cleanup_artifact_id == cleanup_artifact_id))
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=409, detail="Already linked")
    ca = (await db.execute(select(CA).where(CA.id == cleanup_artifact_id))).scalar_one_or_none()
    if not ca:
        raise HTTPException(status_code=404, detail="Cleanup artifact not found")
    if ca.engagement_id != asset.engagement_id:
        raise HTTPException(status_code=400, detail="Cleanup artifact belongs to a different engagement")
    db.add(CleanupArtifactAsset(cleanup_artifact_id=cleanup_artifact_id, asset_id=asset_id))
    await db.commit()


@router.delete("/{asset_id}/cleanup-artifacts/{cleanup_artifact_id}", status_code=status.HTTP_204_NO_CONTENT)
async def unlink_asset_from_cleanup_artifact(asset_id: str, cleanup_artifact_id: str, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    """Remove the link between an asset and a cleanup artifact."""
    from models.associations import CleanupArtifactAsset
    await _require_asset(asset_id, db, current_user)
    result = await db.execute(select(CleanupArtifactAsset).where(CleanupArtifactAsset.asset_id == asset_id, CleanupArtifactAsset.cleanup_artifact_id == cleanup_artifact_id))
    link = result.scalar_one_or_none()
    if not link:
        raise HTTPException(status_code=404, detail="Link not found")
    await db.delete(link)
    await db.commit()
